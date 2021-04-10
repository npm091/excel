import sys
import openpyxl


class ExcellReader:
    """ Excell book reader class """

    def __init__(self, file):
        self.wb = openpyxl.load_workbook(file, data_only=True)

    def getRows(self, sheet, row, count, col):
        list = []
        ws = self.wb[sheet]
        for n in range(row, row+count):
            list.append(ws.cell(row=n, column=col).value)
        return list

    def getCols(self, sheet, row, col, count):
        list = []
        ws = self.wb[sheet]
        for n in range(col, col+count):
            list.append(ws.cell(row=row, column=n).value)
        return list


if __name__ == '__main__':
    args = sys.argv
    if 6 != len(args):
        print("Usage: csvCreator fileName SheetName colNo rowNo count")
        exit(1)
    book = ExcellReader(file=sys.argv[1])
    list = book.getRows(sheet=sys.argv[2], col=int(sys.argv[3]), row=int(sys.argv[4]),
                        count=int(sys.argv[5]))
    comma = ""
    for v in list:
        print(comma + str(v), end="")
        comma = ","

    # ------------------------ test code ---------------------------
    # book = ExcellReader(file="sample.xlsx")
    # list = book.getCols(sheet="data", row=2, col=2, count=5)
    # print(list)
    # list = book.getRows(sheet="data", col=6, row=2, count=50)
    # for v in list:
    #     print(v, end=",")
